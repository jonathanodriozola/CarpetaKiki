import openpyxl
import operator

path = "C:\\Users\\Jonathan\\PycharmProjects\\kikiProject\\Operaciones.xlsx"

def remove(sheet, row):
    for cell in row:
        sheet.delete_rows(row[0].row, 1)

wb_obj = openpyxl.load_workbook(path)

sheet_obj = wb_obj.get_sheet_by_name('1')
CrearHoja = wb_obj.get_sheet_by_name("2")

max_col = sheet_obj.max_column
m_row = sheet_obj.max_row
tupla_excel =()
lista_excel=[]

print(sheet_obj.max_row)

for i in range(2, m_row + 1):
    if(i < sheet_obj.max_row):
        Comitente = sheet_obj.cell(row=i, column=1)
        Fecha = sheet_obj.cell(row=i, column=2)
        TipoMov = sheet_obj.cell(row=i, column=3)
        CodCar = sheet_obj.cell(row=i, column=4)
        TipoCar = sheet_obj.cell(row=i, column=5)
        Especie = sheet_obj.cell(row=i, column=6)
        Cantidad = sheet_obj.cell(row=i, column=7)
        Importe = sheet_obj.cell(row=i, column=8)
        PrecioCom = sheet_obj.cell(row=i, column=9)
        FechaVta = sheet_obj.cell(row=i, column=10)
        CantiVta = sheet_obj.cell(row=i, column=11)
        ImporVta = sheet_obj.cell(row=i, column=12)


        lista_excel.append([
            Comitente.value,
            Fecha.value,
            TipoMov.value,
            CodCar.value,
            TipoCar.value,
            Especie.value,
            Cantidad.value,
            Importe.value,
            PrecioCom.value,
            FechaVta.value,
            CantiVta.value,
            ImporVta.value,
        ])

print("Maximum rows before removing:", sheet_obj.max_row)

#for row in sheet_obj:
#    remove(sheet_obj, row)
#    print("Maximum rows after removing:", sheet_obj.max_row)



PruebaHoja = sorted(lista_excel, key=operator.itemgetter(0,1,2))


for Operaciones in PruebaHoja:
    CrearHoja.append(Operaciones)
wb_obj.save(path)
