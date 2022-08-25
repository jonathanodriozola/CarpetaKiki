import openpyxl


def remove(sheet, row):
    for cell in row:
        sheet.delete_rows(row[0].row, 1)




path = 'C:\\Users\\Jonathan\\PycharmProjects\\kikiProject\\Prueba.xlsx'

book = openpyxl.load_workbook(path)

sheet = book['Hoja1']

print("Maximum rows before removing:", sheet.max_row)

for row in sheet:
    remove(sheet, row)
    print("Maximum rows after removing:", sheet.max_row)



book.save(path)